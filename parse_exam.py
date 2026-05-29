#!/usr/bin/env python3
"""Parse the AI exam HTML using BeautifulSoup and generate Excel."""

import re
import html as html_mod
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

HTML_PATH = "/Users/shisan/.openclaw/media/inbound/AI编程能力认证考试_1---64564459-ed60-475e-852e-ad5d5deb8b46.html"
XLSX_PATH = "/Users/shisan/.openclaw/workspace/AI编程能力认证考试_题目解析.xlsx"

with open(HTML_PATH, 'r', encoding='utf-8') as f:
    soup = BeautifulSoup(f.read(), 'html.parser')

# Find the rvArea div
rv_area = soup.find('div', id='rvArea')
if not rv_area:
    print("ERROR: Could not find rvArea")
    exit(1)

# Find all qcard divs
qcards = rv_area.find_all('div', class_=re.compile(r'qcard'))
print(f"Found {len(qcards)} question cards")

questions = []

for i, qcard in enumerate(qcards):
    try:
        # ---- Extract head info ----
        head_div = qcard.find('div', class_='qcard-head')
        if not head_div:
            continue
        
        spans = head_div.find_all('span')
        if len(spans) < 5:
            continue
        
        qid_tag = spans[0].get_text(strip=True)          # L1-33
        level = spans[1].get_text(strip=True)            # L1/L2/L3
        qtype = spans[2].get_text(strip=True)            # 单选/判断/多选
        score = spans[3].get_text(strip=True)            # 1分/2分
        result_tag = spans[4].get_text(strip=True)       # ✓ 正确
        
        # ---- Extract question text ----
        text_div = qcard.find('div', class_='qcard-text')
        if not text_div:
            continue
        
        full_qtext = text_div.get_text(strip=True)
        # Remove number prefix
        m = re.match(r'(\d+)\.\s*(.*)', full_qtext)
        if m:
            qnum = m.group(1)
            qtext = m.group(2)
        else:
            qnum = str(i + 1)
            qtext = full_qtext
        
        # ---- Extract options ----
        opts_div = qcard.find('div', class_='opts')
        options_data = []
        if opts_div:
            opt_divs = opts_div.find_all('div', class_=re.compile(r'opt'))
            labels = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            
            for idx, opt in enumerate(opt_divs):
                classes = opt.get('class', [])
                is_correct = 'rv-right' in classes
                
                # Get text (remove hidden dcy spans)
                # Remove dcy spans
                for dcy in opt.find_all('span', class_='dcy'):
                    dcy.decompose()
                opt_text = opt.get_text(strip=True)
                
                label = labels[idx] if idx < len(labels) else chr(65 + idx)
                options_data.append({
                    'label': label,
                    'text': opt_text,
                    'is_correct': is_correct
                })
        
        # Correct answers
        correct_answers = [o['label'] for o in options_data if o['is_correct']]
        correct_str = ', '.join(correct_answers) if correct_answers else ''
        
        # Options text for display
        options_text = '\n'.join([f"{o['label']}. {o['text']}" for o in options_data])
        
        # ---- Extract explanation ----
        explain_div = qcard.find('div', class_='qcard-explain')
        explanation = ''
        reference = ''
        if explain_div:
            explain_text_div = explain_div.find('div', class_='qcard-explain-text')
            if explain_text_div:
                explanation = explain_text_div.get_text(strip=True)
            
            source_div = explain_div.find('div', class_='qcard-source')
            if source_div:
                reference = source_div.get_text(strip=True)
        
        # ---- Result status ----
        if 'correct' in result_tag:
            result_status = '正确'
        elif 'partial' in result_tag:
            result_status = '部分正确'
        elif 'wrong' in result_tag:
            result_status = '错误'
        else:
            result_status = result_tag
        
        questions.append({
            'qnum': qnum,
            'qid_tag': qid_tag,
            'level': level,
            'type': qtype,
            'score': score,
            'question_text': qtext,
            'options': options_text,
            'correct_answer': correct_str,
            'result': result_status,
            'explanation': explanation,
            'reference': reference
        })
        
    except Exception as e:
        print(f"Error parsing question {i+1}: {e}")
        import traceback
        traceback.print_exc()
        continue

print(f"Successfully parsed {len(questions)} questions")

# ===== CREATE EXCEL =====
wb = Workbook()

header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='00B4D8', end_color='00B4D8', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

def style_header(ws, headers):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

def style_cell(ws, row, col, value, font=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or Font(name='微软雅黑', size=10)
    cell.alignment = align or Alignment(vertical='top', wrap_text=True)
    cell.border = thin_border
    return cell

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

# ── Sheet 1: 题目列表 ──
ws1 = wb.active
ws1.title = "题目列表"

headers1 = ['序号', '题目编号', '等级', '题型', '分值', '题目内容', '选项', '正确答案', '答题结果', '解析', '参考来源']
style_header(ws1, headers1)

for row_idx, q in enumerate(questions, 2):
    data = [
        q['qnum'], q['qid_tag'], q['level'], q['type'], q['score'],
        q['question_text'], q['options'], q['correct_answer'], q['result'],
        q['explanation'], q['reference']
    ]
    for col_idx, value in enumerate(data, 1):
        cell = style_cell(ws1, row_idx, col_idx, value)
    
    # Color result
    result_cell = ws1.cell(row=row_idx, column=9)
    if '正确' in q['result'] and '部分' not in q['result']:
        result_cell.font = Font(name='微软雅黑', size=10, color='00B050', bold=True)
    elif '部分' in q['result']:
        result_cell.font = Font(name='微软雅黑', size=10, color='FFC000', bold=True)
    elif '错误' in q['result']:
        result_cell.font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)

set_col_widths(ws1, [6, 10, 7, 8, 6, 40, 55, 12, 10, 55, 30])
ws1.freeze_panes = 'A2'

# ── Sheet 2: 答题历史 ──
ws2 = wb.create_sheet("答题历史")

hist_table = soup.find('table', class_='hist-table')
headers2 = ['序号', '类型', '得分', '状态', '提交时间']
style_header(ws2, headers2)

if hist_table:
    rows = hist_table.find_all('tr')[1:]  # skip header
    for row_idx, tr in enumerate(rows, 2):
        cols = tr.find_all('td')
        for col_idx, td in enumerate(cols, 1):
            value = td.get_text(strip=True)
            cell = style_cell(ws2, row_idx, col_idx, value)
            
            # Color status
            if col_idx == 4:
                if '通过' in value:
                    cell.font = Font(name='微软雅黑', size=10, color='00B050', bold=True)
                else:
                    cell.font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
            
set_col_widths(ws2, [8, 10, 12, 10, 22])
ws2.freeze_panes = 'A2'

# ── Sheet 3: 标准答案 ──
ws3 = wb.create_sheet("标准答案")

headers3 = ['题目编号', '等级', '题型', '正确答案', '题目摘要']
style_header(ws3, headers3)

for row_idx, q in enumerate(questions, 2):
    # Summary: first 60 chars
    summary = q['question_text'][:60] + ('...' if len(q['question_text']) > 60 else '')
    data = [q['qid_tag'], q['level'], q['type'], q['correct_answer'], summary]
    for col_idx, value in enumerate(data, 1):
        cell = style_cell(ws3, row_idx, col_idx, value)
    # Bold the answer
    ans_cell = ws3.cell(row=row_idx, column=4)
    ans_cell.font = Font(name='微软雅黑', size=10, color='00B050', bold=True)

set_col_widths(ws3, [12, 8, 10, 15, 55])
ws3.freeze_panes = 'A2'

# ── Sheet 4: 考试信息 ──
ws4 = wb.create_sheet("考试信息")

# Find user info
user_name = ''
user_email = ''
h1_elem = soup.find('b', id='homeUser')
if h1_elem:
    user_name = h1_elem.get_text(strip=True)
email_elem = soup.find('div', id='hiEmail')
if email_elem:
    user_email = email_elem.get_text(strip=True)

# Find score from rvSummary
score_text = ''
rv_summary = soup.find('div', id='rvSummary')
if rv_summary:
    score_text = rv_summary.get_text(strip=True)

ws4.cell(row=1, column=1, value='考试信息').font = Font(name='微软雅黑', bold=True, size=14)
ws4.merge_cells('A1:B1')

info_data = [
    ('考生姓名', user_name),
    ('邮箱', user_email),
    ('考试类型', 'L1-L3（基础级）'),
    ('成绩摘要', score_text),
    ('满分', '100'),
    ('及格线', '90'),
    ('考试系统', 'AI编程能力认证考试系统 · ThunderSoft'),
]

for i, (k, v) in enumerate(info_data, 3):
    style_cell(ws4, i, 1, k, Font(name='微软雅黑', bold=True, size=10))
    style_cell(ws4, i, 2, v)

ws4.column_dimensions['A'].width = 15
ws4.column_dimensions['B'].width = 50

# Save
wb.save(XLSX_PATH)
print(f"\n✅ Excel saved to: {XLSX_PATH}")
print(f"   Sheet 1: 题目列表 ({len(questions)} 题)")
print(f"   Sheet 2: 答题历史")
print(f"   Sheet 3: 标准答案")
print(f"   Sheet 4: 考试信息")
