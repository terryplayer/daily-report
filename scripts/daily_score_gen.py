#!/usr/bin/env python3
"""
每日选股评分自动生成 v2
- 模板指纹检测 + 历史数据存档 + 回溯更新 + 飞书通知
"""

import openpyxl, json, urllib.request, time, sys, os, hashlib
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import Counter

# ========== 配置 ==========
EXCEL_PATH = "/Users/shisan/股票/选股评分报告-完整版.xlsx"
TEMPLATE_SHEET = "选股评分(模版）"
HISTORY_DIR = "/Users/shisan/股票/选股评分历史数据"
LOG_FILE = "/tmp/daily_score_gen.log"
CONFIG_SHEET = "_config"

def log(msg):
    ts = datetime.now().strftime("%H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line)
    with open(LOG_FILE, "a") as f:
        f.write(line + "\n")

def is_trading_day():
    return datetime.now().weekday() < 5

def today_str():
    return datetime.now().strftime("%Y-%m-%d")

def today_name():
    return f"选股评分({today_str()})"

# ========== 模板指纹 ==========
def calc_fingerprint(ws):
    """计算模板sheet的指纹（列数+表头+列宽+关键样式）"""
    fp_data = {
        'cols': ws.max_column,
        'headers': [],
        'widths': [],
        'header_fonts': [],
        'header_fills': [],
    }
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=3, column=col)
        fp_data['headers'].append(str(cell.value or ''))
        fp_data['widths'].append(ws.column_dimensions[get_column_letter(col)].width or 8)
        if cell.font and cell.font.name:
            fp_data['header_fonts'].append(f"{cell.font.name}|{cell.font.size}|{cell.font.bold}|{cell.font.color}")
        if cell.fill and cell.fill.start_color:
            fp_data['header_fills'].append(str(cell.fill.start_color.rgb) if cell.fill.start_color.rgb else '')
    return hashlib.md5(json.dumps(fp_data, sort_keys=True).encode()).hexdigest()

def get_stored_fingerprints(wb):
    """从_config读取所有已存储的指纹"""
    if CONFIG_SHEET not in wb.sheetnames:
        return {}, None
    ws = wb[CONFIG_SHEET]
    stored = {}
    template_fp = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == 'template_fingerprint' and row[1]:
            template_fp = str(row[1])
        elif row[0] and row[1] and row[0] != 'template_updated_at':
            stored[str(row[0])] = str(row[1])
    return stored, template_fp

def save_fingerprint(wb, date_str, fp):
    """保存某日sheet的指纹"""
    ws = _get_config_sheet(wb)
    found = False
    for row in ws.iter_rows(min_row=2, max_col=2):
        if row[0].value == date_str:
            row[1].value = fp
            found = True
            break
    if not found:
        next_row = ws.max_row + 1
        ws.cell(row=next_row, column=1, value=date_str)
        ws.cell(row=next_row, column=2, value=fp)

def _get_config_sheet(wb):
    if CONFIG_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(CONFIG_SHEET)
        ws.sheet_state = 'hidden'
        ws.cell(row=1, column=1, value='key')
        ws.cell(row=1, column=2, value='value')
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
    return wb[CONFIG_SHEET]

# ========== 数据获取 ==========
def fetch_all_stocks(max_time=180):
    """拉取全市场行情，带全局超时保护 max_time 秒。失败则走Tushare备选。"""
    all_data = {}
    total_pages = 56
    start_t = time.time()
    failed_count = 0
    
    for page in range(1, total_pages + 1):
        if time.time() - start_t > max_time:
            log(f"⚠️ 全局超时 {max_time}s，已获取 {len(all_data)} 只，跳过剩余页")
            break
        # 每10页间隔0.5秒，降低限流触发概率
        if page % 10 == 0 and page > 1:
            time.sleep(0.5)
        url = (f'https://push2.eastmoney.com/api/qt/clist/get?'
               f'pn={page}&pz=100&po=1&np=1&fltt=2&invt=2&fid=f3'
               f'&fs=m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23'
               f'&fields=f2,f3,f4,f5,f12,f14,f15,f20,f21,f62,f184,f175,f23')
        success = False
        for retry in range(3):
            if time.time() - start_t > max_time:
                break
            try:
                req = urllib.request.Request(url, headers={'User-Agent':'Mozilla/5.0'})
                resp = urllib.request.urlopen(req, timeout=12).read().decode('utf-8')
                for item in json.loads(resp).get('data',{}).get('diff',[]):
                    all_data[item.get('f12','')] = item
                if page % 10 == 0:
                    log(f"  行情 {page}/{total_pages}页: 累计{len(all_data)}只")
                success = True
                break
            except:
                time.sleep(1 if retry == 0 else 2)
        if not success:
            failed_count += 1
    
    # 如果大量失败（>5页），提前切换到Tushare备选
    if failed_count > 5 or len(all_data) < 3000:
        log(f"⚠️ 东方财富API大量失败({failed_count}页)，切换到Tushare备选...")
        try:
            import tushare as ts
            tk = open('/Users/shisan/.openclaw/workspace/data/tushare_token.txt').read().strip()
            ts.set_token(tk)
            pro = ts.pro_api()
            df = pro.daily(trade_date=datetime.now().strftime("%Y%m%d"))
            if df is not None and len(df) > 1000:
                all_data.clear()
                for _, row in df.iterrows():
                    code = row['ts_code'][:6]
                    all_data[code] = {
                        'f12': code, 'f14': row.get('name', ''), 'f2': row.get('close', 0),
                        'f3': row.get('pct_chg', 0), 'f15': row.get('high', 0), 'f16': row.get('low', 0),
                        'f20': row.get('amount', 0), 'f62': 0
                    }
                log(f"✅ Tushare备选: {len(all_data)}只")
        except Exception as e:
            log(f"❌ Tushare备选也失败: {str(e)[:60]}")
    
    log(f"✅ 行情数据: {len(all_data)}只 (用时{time.time()-start_t:.0f}s)")
    return all_data

def fetch_market_values(all_codes, max_time=60):
    mv = {}
    start_t = time.time()
    total_batches = (len(all_codes) + 80 - 1) // 80
    for i in range(0, len(all_codes), 80):
        if time.time() - start_t > max_time:
            log(f"⚠️ 市值超时 {max_time}s，已获取 {len(mv)} 只")
            break
        batch = all_codes[i:i+80]
        qt = ','.join([f'sh{c}' if c.startswith(('6','9')) else f'sz{c}' for c in batch])
        try:
            req = urllib.request.Request(f'https://qt.gtimg.cn/q={qt}', headers={'User-Agent':'Mozilla/5.0'})
            resp = urllib.request.urlopen(req, timeout=8).read().decode('gbk')
            for line in resp.strip().split(';'):
                if not line.strip(): continue
                p = line.split('~'); c = p[2]
                try:
                    tm = float(p[45]) if len(p)>45 and p[45] else 0
                    fm = float(p[44]) if len(p)>44 and p[44] else 0
                    if tm > 0: mv[c] = {'total_mv': round(tm,2), 'float_mv': round(fm,2)}
                except: pass
        except:
            log(f"  第{i//80+1}批市值请求失败")
        time.sleep(0.1)
        if (i//80+1) % 15 == 0:
            log(f"  市值 {i//80+1}/{total_batches}批: {len(mv)}只")
    for c in all_codes:
        if c not in mv: mv[c] = {'total_mv': 'N/A', 'float_mv': 'N/A'}
    log(f"✅ 市值: {len(mv)}只 (用时{time.time()-start_t:.0f}s)")
    return mv

def save_history(date_str, stocks_data, mv_data):
    path = os.path.join(HISTORY_DIR, date_str)
    os.makedirs(path, exist_ok=True)
    with open(os.path.join(path, "stock_data.json"), "w") as f:
        json.dump(stocks_data, f)
    with open(os.path.join(path, "market_values.json"), "w") as f:
        json.dump(mv_data, f)
    log(f"💾 历史数据已存档: {path}")

def load_history(date_str):
    path = os.path.join(HISTORY_DIR, date_str)
    if not os.path.exists(path):
        return None, None
    try:
        with open(os.path.join(path, "stock_data.json")) as f:
            sd = json.load(f)
        with open(os.path.join(path, "market_values.json")) as f:
            mv = json.load(f)
        return sd, mv
    except:
        return None, None

# ========== 分类和评分 ==========
SECTOR_RULES = [
    ("科技/半导体", ['芯片','半导体','集成电路','电子','软件','信息','通信','IT','数据','智能',
                    '计算','网络','互联','微','光刻','AI','人工','机器','自动化','数码','科技','华强']),
    ("通信/电子", ['通信','电子','光电','光学','元件','器件','显示','面板','LED','传感','射频']),
    ("AI/数字经济", ['传媒','游戏','互联网','数字','AI','人工智能','娱乐','广告','营销','影视',
                     '文化','教育','出版','直播','短视频','东方财富','完美','蓝色']),
    ("化工/材料", ['化工','化学','材料','金属','钢铁','锂','钴','镍','稀土','有色','合金',
                     '化纤','塑料','橡胶','石化','农药','化肥','纤维','玻纤','碳']),
    ("能源/公用事业", ['电力','能源','燃气','水务','环保','新能源','光伏','风电','核电',
                       '水电','火电','电网','节能','环境','水','西电']),
]

def market_type(code, name):
    if '退' in str(name): return '退市'
    if str(code).startswith(('4','8','92')): return '北交所'
    if str(code).startswith('688'): return '科创板'
    if str(code).startswith('6'): return '沪市主板'
    if str(code).startswith(('300','301')): return '创业板'
    if str(code).startswith(('00','001','002','003')): return '深市主板'
    return '其他'

def classify_sector(name):
    for sec, kws in SECTOR_RULES:
        for kw in kws:
            if kw in str(name): return sec
    return "其他"

def calc_scores(stocks_data, mv_data):
    scores = []
    for code, item in stocks_data.items():
        name = str(item.get('f14', ''))
        def sf(v):
            try: return float(v) if v else 0
            except: return 0
        scores.append({
            'code': code, 'name': name,
            'sector': classify_sector(name),
            'market': market_type(code, name),
            'price': sf(item.get('f2')), 'chg': sf(item.get('f3')),
            'chg_amt': sf(item.get('f4')),
            'high': sf(item.get('f15')), 'low': sf(item.get('f5')),
            'turn': sf(item.get('f20')), 'vr': sf(item.get('f21')),
            'amt': sf(item.get('f21')),
            'pe': sf(item.get('f175')), 'pb': sf(item.get('f23')),
            'total_mv': mv_data.get(code, {}).get('total_mv', '—'),
            'fund_inflow': round(sf(item.get('f62'))/1e8, 2) if item.get('f62') else '—',
            'roe': sf(item.get('f184')),
            'rs': 0, 'multi': 0, 'rank': 0
        })
    if not scores: return scores
    chg = [s['chg'] for s in scores]; cs = sorted(chg); n = len(cs)
    tr = [s['turn'] for s in scores]; vr = [s['vr'] for s in scores]
    pei = [1/s['pe'] if s['pe'] and s['pe']>0 else 0 for s in scores]
    pbi = [1/s['pb'] if s['pb'] and s['pb']>0 else 0 for s in scores]
    def pct(vals, t, hi=True):
        if not vals: return 50
        return sum(1 for v in vals if (v<=t if hi else v>=t))/len(vals)*100
    for s in scores:
        s['rs'] = round(sum(1 for v in cs if v<=s['chg'])/n*100, 1)
        c = pct(chg, s['chg'], True); t = pct(tr, s['turn'], True); v = pct(vr, s['vr'], True)
        p1 = 1/s['pe'] if s['pe'] and s['pe']>0 else 0
        p2 = 1/s['pb'] if s['pb'] and s['pb']>0 else 0
        s['multi'] = round(c*0.30 + t*0.20 + v*0.20 + pct(pei,p1,True)*0.15 + pct(pbi,p2,True)*0.15, 1)
    scores.sort(key=lambda x: x['multi'], reverse=True)
    for i, s in enumerate(scores, 1): s['rank'] = i
    return scores

def rg(score):
    if score>=90: return 'A+'
    if score>=75: return 'A'
    if score>=55: return 'B'
    if score>=35: return 'C'
    return 'D'

# ========== 填充Excel ==========
def fill_sheet(ws, scores, date_str, is_template=False):
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        if val: headers[str(val).strip()] = col
    log(f"   表头: {list(headers.keys())}")

    # 清空旧数据
    for r in range(4, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    field_map = {
        '排名':'rank','代码':'code','名称':'name','最新价':'price','涨跌幅%':'chg',
        '涨跌额':'chg_amt','最高':'high','最低':'low','换手率%':'turn','量比':'vr',
        '市盈率':'pe','市净率':'pb','成交额(亿)':'amt','RS得分':'rs','多因子':'multi',
        '等级':'grade','板块':'sector','总市值(亿)':'total_mv','ROE%':'roe',
        '主力净流入(亿)':'fund_inflow','市场':'market'
    }
    def get_val(s, field):
        if field == 'grade': return rg(s['rs'])
        if field in ('pe','pb'): return s[field] if s[field]!=0 else '—'
        v = s.get(field,'—')
        return round(v,2) if isinstance(v,float) else (v if v!=0 else 0)
    
    for idx, s in enumerate(scores):
        r = idx + 4
        for h, col in headers.items():
            f = field_map.get(h)
            if f: ws.cell(row=r, column=col, value=get_val(s, f))
    
    # 更新标题日期
    for c in range(1, ws.max_column+1):
        cell = ws.cell(row=1, column=c)
        if cell.value: cell.value = str(cell.value).replace('2026-06-01', date_str)
        cell = ws.cell(row=2, column=c)
        if cell.value: cell.value = str(cell.value).replace('2026-06-01', date_str)
    
    # 更新筛选范围
    ws.auto_filter.ref = f'A3:{get_column_letter(max(headers.values()))}{len(scores)+3}'
    log(f"   ✅ 填充: {len(scores)}行")

# ========== 市场全景HTML ==========
def gen_selestock_html(scores, date_str):
    import json
    hdrs = ['排名','代码','名称','持仓','最新价','涨跌幅%','涨跌额','最高','最低',
            '换手率%','量比','市盈率','市净率','成交额(亿)','RS得分','多因子','等级',
            '板块','总市值(亿)','ROE%','主力净流入(亿)','市场']
    data = []
    for s in scores:
        item = {}
        for h in hdrs:
            if h == '排名': item[h] = s['rank']
            elif h == '代码': item[h] = s['code']
            elif h == '名称': item[h] = s['name']
            elif h == '等级': item[h] = rg(s['rs'])
            elif h == '板块': item[h] = s['sector']
            elif h == '市场': item[h] = s['market']
            elif h == '总市值(亿)': item[h] = s['total_mv']
            elif h == 'ROE%': item[h] = s['roe']
            elif h == '主力净流入(亿)': item[h] = s['fund_inflow']
            elif h == '成交额(亿)': item[h] = round(s['amt']/1e8,2) if s['amt'] else 0
            elif h == '市盈率': item[h] = s['pe'] if s['pe']!=0 else '—'
            elif h == '市净率': item[h] = s['pb'] if s['pb']!=0 else '—'
            else: item[h] = s.get({'最新价':'price','涨跌幅%':'chg','涨跌额':'chg_amt','最高':'high','最低':'low',
                                   '换手率%':'turn','量比':'vr','RS得分':'rs','多因子':'multi'}.get(h,''), '—')
        data.append(item)
    
    hdrs_j = json.dumps(hdrs, ensure_ascii=False)
    data_j = json.dumps(data, ensure_ascii=False, default=str)
    total = len(data)
    
    html = f'''<!DOCTYPE html><html lang="zh-CN"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"><title>📊 市场全景 | {date_str}</title><style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:-apple-system,"PingFang SC","Microsoft YaHei",sans-serif;background:#0d1117;color:#c9d1d9;padding:20px}}
.container{{max-width:1400px;margin:0 auto}}
h1{{color:#58a6ff;font-size:22px;margin-bottom:4px}}
.sub{{color:#8b949e;font-size:13px;margin-bottom:16px}}
.ctl{{display:flex;gap:12px;margin-bottom:12px;flex-wrap:wrap;align-items:center}}
.ctl select,.ctl input{{background:#161b22;color:#c9d1d9;border:1px solid #30363d;padding:6px 12px;border-radius:6px;font-size:13px;outline:none}}
.ctl select{{cursor:pointer}}.ctl input{{flex:1;min-width:150px}}
.tw{{overflow-x:auto;border:1px solid #30363d;border-radius:8px;background:#161b22}}
table{{width:100%;border-collapse:collapse;font-size:12px;white-space:nowrap}}
th{{background:#1c2333;color:#8b949e;font-weight:600;padding:8px 10px;text-align:center;border-bottom:1px solid #30363d;cursor:pointer;position:sticky;top:0;z-index:1}}
th:hover{{color:#58a6ff}}
td{{padding:5px 8px;text-align:center;border-bottom:1px solid #21262d}}
tr:hover td{{background:#1a2744}}
.up{{color:#f85149}}.dn{{color:#3fb950}}.gd{{color:#d9a52e}}.pu{{color:#bf77f6}}.bl{{color:#58a6ff}}.gy{{color:#8b949e}}.gn{{color:#3fb950}}
.r1{{color:#f85149;font-weight:700}}.r2{{color:#d9a52e;font-weight:700}}.r3{{color:#58a6ff;font-weight:700}}
.tag{{display:inline-block;padding:1px 6px;border-radius:4px;font-size:11px;font-weight:600}}
.t-ap{{background:#f8514922;color:#f85149}}.t-a{{background:#d9a52e22;color:#d9a52e}}
.t-b{{background:#58a6ff22;color:#58a6ff}}.t-c{{background:#8b949e22;color:#8b949e}}.t-d{{background:#3fb95022;color:#3fb950}}
.zg{{background:#3fb95011}}.zy{{background:#d9a52e11}}.zr{{background:#f8514911}}
.pag{{display:flex;justify-content:center;gap:6px;margin-top:12px;flex-wrap:wrap}}
.pag button{{background:#161b22;color:#c9d1d9;border:1px solid #30363d;padding:6px 12px;border-radius:6px;cursor:pointer;font-size:13px}}
.pag button:hover{{background:#1c2333;border-color:#58a6ff}}
.pag .act{{background:#1f6feb;border-color:#1f6feb}}.cnt{{color:#8b949e;font-size:13px}}
.ft{{text-align:center;color:#8b949e;font-size:12px;margin-top:20px;padding-top:12px;border-top:1px solid #21262d}}
.wb{{display:inline-block;padding:1px 8px;border-radius:4px;font-size:11px;font-weight:600;cursor:pointer;border:none}}
.wb-hld{{background:#3fb95022;color:#3fb950;cursor:default}}
.wb-un{{background:#8b949e22;color:#8b949e}}
.wb-add{{background:#58a6ff22;color:#58a6ff}}
.wb-add:hover{{background:#58a6ff44}}
.wb-rm{{background:#f8514922;color:#f85149}}
.wb-rm:hover{{background:#f8514944}}
.wb-loading{{opacity:0.5;pointer-events:none}}
.fl-btn{{display:inline-block;padding:4px 12px;border-radius:12px;font-size:11px;font-weight:600;cursor:pointer;border:1px solid #30363d;background:#0d1117;color:#8b949e;margin-right:4px;user-select:none}}
.fl-btn:hover{{border-color:#58a6ff;color:#e6edf3}}
.fl-btn.act{{background:#58a6ff;color:#fff;border-color:#58a6ff}}
</style></head><body><div class=container>
<h1>📊 市场全景 | {date_str}</h1>
<p class=sub>共{total}只 · 双击列头排序</p>
<div class=ctl>
<select id=sf onchange=r()><option value=>全部板块</option><option>科技/半导体</option><option>通信/电子</option><option>AI/数字经济</option><option>化工/材料</option><option>能源/公用事业</option><option>其他</option></select>
<span class=fl-btn act data-fw=all onclick=fw(this,'all')>📋 全部</span>
<span class=fl-btn data-fw=hold onclick=fw(this,'hold')>🔴 仅持仓</span>
<span class=fl-btn data-fw=unhold onclick=fw(this,'unhold')>⚪ 仅未持仓</span>
<select id=mf onchange=r()><option value=>全部市场</option><option>沪市主板</option><option>深市主板</option><option>创业板</option><option>科创板</option><option>北交所</option><option>退市</option></select>
<select id=gf onchange=r()><option value=>全部等级</option><option>A+</option><option>A</option><option>B</option><option>C</option><option>D</option></select>
<input id=si placeholder="🔍 代码/名称..." oninput=r() style=flex:1>
<span class=cnt id=cn>0条</span></div>
<div class=tw><table><thead><tr>'''
    for h in hdrs:
        html += f'<th onclick=st("{h}")>{h}</th>'
    html += '</tr></thead><tbody id=tb></tbody></table></div>'
    html += '<div class=pag id=pg></div><p class=ft>🌀 市场全景 · 数据仅供参考</p></div>'
    html += f'<script>const PS=50,ad={data_j},hd={hdrs_j};'
    html += '''let cp=1,sk=null,sa=!1;
function gv(it,k){let v=it[k];if(v==='—'||v==null)return '';if(typeof v==='number')return v;return isNaN(v=parseFloat(v))?v:v}
function fl(){let s=document.getElementById('sf').value,m=document.getElementById('mf').value,g=document.getElementById('gf').value,q=document.getElementById('si').value.toLowerCase()
return ad.filter(it=>{if(s&&it['板块']!==s)return 0;if(m&&it['市场']!==m)return 0;if(g&&it['等级']!==g)return 0;if(q&&!String(it['代码']).includes(q)&&!String(it['名称']).toLowerCase().includes(q))return 0;return 1})}
function st(k){sk===k?sa=!sa:(sk=k,sa=!1);r()}
let wh={};let wf='all';
function fw(el,m){wf=m;document.querySelectorAll('.fl-btn').forEach(b=>b.classList.toggle('act',b===el));r()}
async function loadWatchlist(){try{let r=await fetch('http://127.0.0.1:18790/watchlist');let d=await r.json();d.codes.forEach(c=>wh[c]=1)}catch(e){console.log('⚠️ 未连接持仓API')}}
function cl(k,v,code){
if(v===''||v==null)return '<td class=gy>—</td>'
if(k==='排名')return '<td class='+(v===1?'r1':v===2?'r2':v===3?'r3':'')+'>'+v+'</td>'
if(k==='持仓'){let isH=wh[code];return isH?'<td><span class="wb wb-hld">🔴 已持仓</span><span class="wb wb-rm" onclick=toggleHold("'+code+'","'+v+'",0)>➖</span></td>':'<td><span class="wb wb-un">⚪ 未持仓</span><span class="wb wb-add" onclick=toggleHold("'+code+'","'+v+'",1)>➕</span></td>'}
if(k==='涨跌幅%'||k==='涨跌额'){let n=parseFloat(v);if(isNaN(n))return '<td>'+v+'</td>';return '<td class='+(n>0?'up':n<0?'dn':'gd')+'>'+(n>0?'+':'')+v+(k==='涨跌幅%'?'%':'')+'</td>'}
if(k==='等级'){return '<td><span class=tag t-'+String(v).toLowerCase().replace('+','ap')+'>'+v+'</span></td>'}
if(k==='板块'){return '<td class='+({'科技/半导体':'up','通信/电子':'gd','AI/数字经济':'pu','化工/材料':'bl','能源/公用事业':'gn','其他':'gy'}[String(v)]||'')+'>'+v+'</td>'}
if(k==='市场'){return '<td class='+({'沪市主板':'bl','深市主板':'gn','创业板':'pu','科创板':'gd','北交所':'dn','退市':'up'}[String(v)]||'')+'>'+v+'</td>'}
if(['换手率%','量比','RS得分','多因子'].includes(k)){let n=parseFloat(v);if(!isNaN(n)){return '<td class='+(n<=40||k==='换手率%'&&n<=3||k==='量比'&&n<=0.8?'zg':n<=70||k==='换手率%'&&n<=10||k==='量比'&&n<=1.5?'zy':'zr')+'>'+v+'</td>'}}
if(k==='市盈率'){let n=parseFloat(v);if(!isNaN(n)&&n>0)return '<td class='+(n<=15?'zg':n<=30?'zy':'zr')+'>'+v+'</td>'}
if(k==='主力净流入(亿)'){let n=parseFloat(v);if(!isNaN(n))return '<td class='+(n>0.5?'up':n<-0.5?'dn':'')+'>'+v+'</td>'}
return '<td class=gd>'+(typeof v==='number'?v.toFixed(2):v)+'</td>'}
async function toggleHold(code,name,add){
let btns=document.querySelectorAll('.wb');btns.forEach(b=>b.classList.add('wb-loading'))
try{let url='http://127.0.0.1:18790/watchlist/'+(add?'add?code='+code+'&name='+encodeURIComponent(name):'remove?code='+code);await fetch(url,{method:add?'POST':'DELETE'});if(add)wh[code]=1;else delete wh[code];r()}catch(e){alert('操作失败: '+e.message)}
btns.forEach(b=>b.classList.remove('wb-loading'))}
function r(){let f=fl();
if(wf==='hold')f=f.filter(it=>wh[it['代码']]);
if(wf==='unhold')f=f.filter(it=>!wh[it['代码']]);
if(sk)f.sort((a,b)=>{let va=gv(a,sk),vb=gv(b,sk);return typeof va==='number'&&typeof vb==='number'?(sa?va-vb:vb-va):sa?String(va).localeCompare(String(vb)):String(vb).localeCompare(String(va))})
document.getElementById('cn').textContent=f.length+'条'
let tp=Math.ceil(f.length/PS)||1;if(cp>tp)cp=tp;let s=(cp-1)*PS,pg=f.slice(s,s+PS)
document.getElementById('tb').innerHTML=pg.map(it=>'<tr>'+hd.map(k=>cl(k,it[k],it['代码'])).join('')+'</tr>').join('')
let pb='',st=Math.max(1,cp-7),en=Math.min(tp,st+15);if(st>1)pb+='<button onclick=gp('+(cp-1)+')>◀</button>'
for(let i=st;i<=en;i++)pb+='<button class='+(i===cp?'act':'')+' onclick=gp('+i+')>'+i+'</button>'
if(en<tp)pb+='<button onclick=gp('+Math.min(cp+1,tp)+')>▶</button>'
document.getElementById('pg').innerHTML=pb}
function gp(p){cp=p;r()}

loadWatchlist();setTimeout(r,100)
</script></body></html>'''

    # 保存两份：selestock/index.html（最新版）+ selestock/YYYY-MM-DD.html（历史版）
    sel_dir = os.path.join('/Users/shisan/.openclaw/workspace', 'daily-report-html', 'selestock')
    if not os.path.exists(sel_dir):
        os.makedirs(sel_dir, exist_ok=True)
    
    with open(os.path.join(sel_dir, 'index.html'), 'w') as f:
        f.write(html)
    with open(os.path.join(sel_dir, f'{date_str}.html'), 'w') as f:
        f.write(html)
    log(f"✅ 市场全景HTML: selestock/ ({total}只)")

# ========== 主流程 ==========
def main():
    log("=" * 50)
    log("📊 每日选股评分 v2 开始")
    
    if not is_trading_day():
        log("⏭️ 非开盘日跳过")
        return "⏭️ 非开盘日，跳过"

    date_str = today_str()
    sheet_name = today_name()
    log(f"📅 {date_str}")
    
    if not os.path.exists(EXCEL_PATH):
        log(f"❌ 文件不存在: {EXCEL_PATH}")
        return "❌ Excel文件不存在"
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    if TEMPLATE_SHEET not in wb.sheetnames:
        log(f"❌ 无模板sheet")
        wb.close()
        return "❌ 模板sheet不存在"
    
    template_ws = wb[TEMPLATE_SHEET]
    current_fp = calc_fingerprint(template_ws)
    stored_fps, old_fp = get_stored_fingerprints(wb)
    template_changed = (old_fp is not None and old_fp != current_fp)
    
    if template_changed:
        log(f"🔄 模板已变更! (old={old_fp[:8]} new={current_fp[:8]})")
    else:
        log(f"✅ 模板未变化 (fp={current_fp[:8]})")
    
    # 保存模板指纹
    cws = _get_config_sheet(wb)
    found = False
    for row in cws.iter_rows(min_row=2, max_col=2):
        if row[0].value == 'template_fingerprint':
            row[1].value = current_fp
            found = True; break
    if not found:
        cws.cell(row=2, column=1, value='template_fingerprint')
        cws.cell(row=2, column=2, value=current_fp)
    cws.cell(row=3, column=1, value='template_updated_at')
    cws.cell(row=3, column=2, value=datetime.now().isoformat())
    
    # 确定需要生成的sheet列表
    need_generate = []
    
    # A) 回溯：模板变了 → 重新生成所有现有sheet
    if template_changed:
        for sn in wb.sheetnames:
            if sn.startswith("选股评分(20") and sn != sheet_name:
                d = sn.replace("选股评分(","").replace(")","")
                stored_fp = stored_fps.get(d)
                if stored_fp != current_fp:
                    need_generate.append((sn, d))
        if need_generate:
            log(f"🔄 需回溯 {len(need_generate)} 个sheet")
    
    # B) 今日sheet不存在 → 需要新生成
    need_new_today = sheet_name not in wb.sheetnames
    
    # 获取/加载今日数据
    today_stocks, today_mv = None, None
    
    # 市场全景需要全量数据，即使Sheet已存在也要拉取
    log("🔄 获取今日行情数据（用于市场全景）...")
    today_stocks = fetch_all_stocks()
    if today_stocks:
        # 数据完整性检查：全市场正常应≥5000只，不足则用历史存档兜底
        if len(today_stocks) < 5000:
            log(f"⚠️ 数据不完整({len(today_stocks)}只)，尝试加载历史存档兜底...")
            h_sd, h_mv = load_history(date_str)
            if h_sd and len(h_sd) >= len(today_stocks):
                today_stocks, today_mv = h_sd, h_mv
                log(f"✅ 使用历史存档: {len(today_stocks)}只")
            else:
                today_mv = fetch_market_values(list(today_stocks.keys()))
                save_history(date_str, today_stocks, today_mv)
                log(f"✅ 行情数据: {len(today_stocks)}只（不全，但无更完整存档）")
        else:
            today_mv = fetch_market_values(list(today_stocks.keys()))
            save_history(date_str, today_stocks, today_mv)
            log(f"✅ 行情数据: {len(today_stocks)}只")
    elif not need_new_today:
        # 拉取失败，尝试加载历史
        log("⚠️ 拉取失败，尝试加载历史数据...")
        today_stocks, today_mv = load_history(date_str)
        if today_stocks:
            log(f"✅ 加载历史数据: {len(today_stocks)}只")
    
    # 生成今日sheet
    scores = None
    if need_new_today and today_stocks:
        scores = calc_scores(today_stocks, today_mv)
        new_ws = wb.copy_worksheet(template_ws)
        new_ws.title = sheet_name
    elif today_stocks and today_mv:
        # Sheet已存在，只需计算评分用于市场全景
        scores = calc_scores(today_stocks, today_mv)
        log(f"✅ 评分已计算: {len(scores)}只（用于市场全景）")
    
    # 回溯重新生成
    for sn, d in need_generate:
        log(f"🔄 回溯: {sn} ({d})...")
        h_sd, h_mv = load_history(d)
        if h_sd and h_mv:
            scores = calc_scores(h_sd, h_mv)
            # 删除旧的sheet
            old_idx = wb.sheetnames.index(sn)
            old_sheet = wb[sn]
            wb.remove(old_sheet)
            # 从模板重新创建
            new_ws = wb.copy_worksheet(template_ws)
            new_ws.title = sn
            fill_sheet(new_ws, scores, d)
            # 移到原来位置
            new_idx = wb.sheetnames.index(sn)
            target = min(old_idx, len(wb.sheetnames)-1)
            wb.move_sheet(sn, offset=target-new_idx)
            save_fingerprint(wb, d, current_fp)
            log(f"   ✅ 回溯完成: {sn}")
        else:
            log(f"   ⚠️ 无历史数据，跳过: {sn}")
    
    # 保存
    log("💾 保存文件...")
    wb.save(EXCEL_PATH)
    wb.close()
    
    mkt_dist = Counter()
    if today_stocks:
        for item in today_stocks.values():
            mkt_dist[market_type(str(item.get('f12','')), str(item.get('f14','')))] += 1
    
    result = f"✅ 完成"
    if need_new_today: result += f" | 新增: {sheet_name}"
    if need_generate: result += f" | 回溯: {len(need_generate)}个"
    
    # 生成市场全景HTML
    if scores:
        try:
            gen_selestock_html(scores, date_str)
            result += " | 🌐 市场全景已生成"
        except Exception as e:
            log(f"⚠️ 市场全景生成失败: {e}")
    
    log(result)
    return result

if __name__ == "__main__":
    result = main()
    # 输出结果（cron会捕获并推送）
    print(f"\n[RESULT] {result}")
