#!/usr/bin/env python3
"""
📊 Layer 4: 评分计算引擎
从 data/raw/ 读取原始数据 → 计算 RS/多因子/多因素 → 输出 Excel
独立运行，不依赖数据采集层
"""

import json, os, sys, math, time
from datetime import date

WORKSPACE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR = os.path.join(WORKSPACE, "data", "raw")
SCORES_DIR = os.path.join(WORKSPACE, "data", "scores")
CONFIG_PATH = os.path.join(WORKSPACE, "data", "config", "fields.json")
WATCH_PATH = os.path.join(WORKSPACE, "scripts", "template-stocks.json")
os.makedirs(SCORES_DIR, exist_ok=True)

def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}")

# ─── 加载配置 ───────────────────────────────────
def load_config():
    with open(CONFIG_PATH) as f:
        return json.load(f)

def load_watchlist():
    with open(WATCH_PATH) as f:
        d = json.load(f)
    return {s["code"]: s for s in d.get("watchlist", [])}

def load_raw(date_str):
    path = os.path.join(RAW_DIR, f"{date_str}.json")
    if not os.path.exists(path):
        log(f"  ❌ {date_str} 数据不存在")
        return None
    with open(path) as f:
        return json.load(f)

# ─── 评分计算 ───────────────────────────────────
def calc_rs(stocks, watch_codes):
    """RS排名：基于涨跌幅的截面相对强度"""
    # 只算持仓股票
    hold_stocks = [(code, s) for code, s in stocks.items() if code in watch_codes]
    if not hold_stocks:
        log("  ⚠️ 持仓股票无数据")
        return {}
    
    sorted_by_chg = sorted(hold_stocks, key=lambda x: x[1].get("f3", 0) or 0, reverse=True)
    total = len(sorted_by_chg)
    
    rs = {}
    for idx, (code, s) in enumerate(sorted_by_chg):
        # RS 分数：0~100，排名越靠前越高
        score = round((1 - idx / total) * 100, 1)
        chg = s.get("f3", 0) or 0
        # 等级
        if score >= 80: grade = "A+"
        elif score >= 60: grade = "A"
        elif score >= 40: grade = "B"
        elif score >= 20: grade = "C"
        else: grade = "D"
        
        rs[code] = {
            "rank": idx + 1, "total": total,
            "rs_score": score, "grade": grade,
            "chg": chg,
        }
    
    log(f"  ✅ RS: {total}只")
    return rs

def calc_multi_factor(stocks, watch_codes, rs):
    """多因子评分"""
    if not rs:
        return {}
    
    scores = []
    for code in watch_codes:
        if code not in stocks:
            continue
        s = stocks[code]
        r = rs.get(code, {})
        
        # 因子1：RS (25%)
        rs_score = r.get("rs_score", 50) if r else 50
        
        # 因子2：涨跌幅作为动量代理 (20%)
        chg = s.get("f3", 0) or 0
        mom = max(0, min(100, 50 + chg * 5))
        
        # 因子3：换手率作为偏离度代理 (10%)
        turn = s.get("f62", 0) or 0
        mrd = max(0, min(100, turn * 3)) if turn > 0 else 50
        
        # 因子4：多因子综合 (30%) — 基于 PE/PB/换手率/量比
        pe = s.get("f175", 0) or 0
        pb = s.get("f23", 0) or 0
        vr = s.get("f168", 0) or 0
        score_pe = max(0, min(100, 50 + (30 - pe) * 1.5)) if pe > 0 else 50
        score_pb = max(0, min(100, 50 + (5 - pb) * 10)) if pb > 0 else 50
        score_vr = max(0, min(100, vr * 50)) if vr > 0 else 50
        score_turn = max(0, min(100, turn * 5)) if turn > 0 else 50
        multi = round((score_pe * 0.25 + score_pb * 0.25 + score_vr * 0.25 + score_turn * 0.25), 1)
        
        # 因子5：行业轮动 (15%) — 简化为板块均值
        rotation = 50
        
        # 综合
        total = round(
            rs_score * 0.25 +
            mom * 0.20 +
            mrd * 0.10 +
            multi * 0.30 +
            rotation * 0.15,
            1
        )
        
        scores.append({
            "code": code,
            "name": s.get("f14", ""),
            "sector": s.get("f100") or "其他",
            "price": s.get("f2"),
            "chg": chg,
            "turn": turn,
            "pe": pe,
            "pb": pb,
            "rs_score": rs_score,
            "rs_rank": r.get("rank", 0),
            "grade": r.get("grade", "C"),
            "mom": round(mom, 1),
            "mrd": round(mrd, 1),
            "multi": multi,
            "rotation": rotation,
            "total_score": total,
        })
    
    # 按综合评分排序
    scores.sort(key=lambda x: x["total_score"], reverse=True)
    for idx, s in enumerate(scores):
        s["rank"] = idx + 1
    
    log(f"  ✅ 多因子: {len(scores)}只")
    return scores

def output_excel(scores, date_str):
    """输出 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        log("  ❌ 需要 openpyxl: pip3 install openpyxl")
        return None
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"选股评分({date_str})"
    
    # 表头
    headers = ["排名","代码","名称","最新价","涨跌幅%","换手率%",
               "市盈率","市净率","RS得分","RS等级","动量","偏离度",
               "多因子","轮动","综合评分","板块"]
    hf = Font(bold=True, size=10, color="FFFFFF")
    hfill = PatternFill(start_color="1F6FEB", end_color="1F6FEB", fill_type="solid")
    thin = Border(*(Side(style="thin", color="D0D0D0"),) * 4)
    
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hf
        c.fill = hfill
        c.alignment = Alignment(horizontal="center")
        c.border = thin
    
    ws.cell(row=1, column=len(headers)+1, value=f"生成时间: {time.strftime('%Y-%m-%d %H:%M')}").font = Font(size=9, color="666666")
    
    # 数据
    for idx, s in enumerate(scores):
        r = idx + 2
        vals = [
            s["rank"], s["code"], s["name"], s["price"], s["chg"], s["turn"],
            s["pe"], s["pb"], s["rs_score"], s["grade"],
            s["mom"], s["mrd"], s["multi"], s["rotation"],
            s["total_score"], s["sector"]
        ]
        for col, v in enumerate(vals, 1):
            ws.cell(row=r, column=col, value=v).border = thin
    
    path = os.path.join(SCORES_DIR, f"{date_str}.xlsx")
    wb.save(path)
    log(f"  ✅ Excel: {path} ({os.path.getsize(path)/1024:.0f}KB)")
    return path

def output_market_panorama(scores, watch_codes, date_str):
    """输出评分后的市场全景数据（仅持仓）"""
    pass  # 后续扩展

def main():
    date_str = date.today().strftime("%Y-%m-%d")
    
    # 参数
    if len(sys.argv) > 1:
        date_str = sys.argv[1]
    
    log(f"📊 评分计算开始: {date_str}")
    
    cfg = load_config()
    log(f"  配置: RS{cfg['score']['rs_method']} 权重={cfg['score']['weights']}")
    
    watch = load_watchlist()
    watch_codes = set(watch.keys())
    log(f"  持仓: {len(watch_codes)}只")
    
    raw = load_raw(date_str)
    if not raw:
        return
    
    meta = raw.pop("_meta", {})
    log(f"  原始数据: {len(raw)}只 (来源: {meta.get('source', '?')})")
    
    # 计算 RS
    log(f"  📐 计算 RS...")
    rs = calc_rs(raw, watch_codes)
    if not rs:
        log(f"  ❌ RS 计算失败，无持仓数据")
        return
    
    # 计算多因子
    log(f"  📐 计算多因子评分...")
    scores = calc_multi_factor(raw, watch_codes, rs)
    
    # 输出
    log(f"  💾 生成 Excel...")
    path = output_excel(scores, date_str)
    
    if path:
        print(f"\n[RESULT] ✅ 评分完成: {date_str} | {len(scores)}只 | {path}")

if __name__ == "__main__":
    main()
