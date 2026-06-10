#!/usr/bin/env python3
"""
🚀 Layer 2: 数据服务 (端口 18790)
提供：原始数据查询/修改 + 评分触发 + 持仓管理
纯标准库实现，零依赖
"""

import json, os, sys, time, shutil
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
from datetime import date, datetime

WORKSPACE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RAW_DIR = os.path.join(WORKSPACE, "data", "raw")
SCORES_DIR = os.path.join(WORKSPACE, "data", "scores")
WATCH_CONFIG = os.path.join(WORKSPACE, "config", "stocks.json")
TEMPLATE_CONFIG = os.path.join(WORKSPACE, "scripts", "template-stocks.json")
FIELD_DEFS = {
    "f2": "最新价(元)", "f3": "涨跌幅(%)", "f12": "代码", "f14": "名称",
    "f15": "最高价(元)", "f16": "最低价(元)", "f17": "今开(元)", "f18": "昨收(元)",
    "f20": "成交额(亿元)", "f21": "流通市值(亿元)", "f23": "市净率",
    "f62": "换手率(%)", "f115": "市盈率(动)", "f168": "量比",
    "f175": "市盈率(静)", "f184": "ROE(%)", "f100": "细分行业", "f193": "主力净流入(万)"
}

# ─── 数据存取 ───────────────────────────────────
def load_raw(date_str=None):
    if not date_str:
        date_str = date.today().strftime("%Y-%m-%d")
    path = os.path.join(RAW_DIR, f"{date_str}.json")
    if not os.path.exists(path):
        # 降级：自动找最近可用数据
        files = sorted([f for f in os.listdir(RAW_DIR) if f.endswith('.json')], reverse=True)
        if files:
            fallback = files[0]
            path = os.path.join(RAW_DIR, fallback)
            with open(path) as f:
                data = json.load(f)
            fallback_date = fallback.replace('.json', '')
            # 在 _meta 中标记降级信息
            if '_meta' not in data:
                data['_meta'] = {}
            data['_meta']['is_fallback'] = True
            data['_meta']['latest_date'] = fallback_date
            data['_meta']['fallback_hint'] = f"今日数据19:00采集后更新，当前显示{fallback_date}数据"
            return data, None
        return None, f"无 {date_str} 数据，且无历史数据可降级"
    with open(path) as f:
        return json.load(f), None

def save_raw(date_str, data):
    path = os.path.join(RAW_DIR, f"{date_str}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def load_watchlist():
    with open(WATCH_CONFIG) as f:
        return json.load(f)

def save_watchlist(data):
    with open(WATCH_CONFIG, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    with open(TEMPLATE_CONFIG, "w") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_sector(code):
    if code.startswith("688"):  return "科技/半导体"
    if code.startswith(("300","301")):  return "创业板/科技"
    if code.startswith("6"):   return "主板"
    if code.startswith(("0","2")): return "深市主板"
    if code.startswith(("8","4")): return "北交所"
    return "其他"

# ─── HTTP Handler ───────────────────────────────
class NanToNullEncoder(json.JSONEncoder):
    def default(self, obj):
        return None
    def encode(self, o):
        return super().encode(self._clean(o))
    def _clean(self, o):
        import math
        if isinstance(o, float) and math.isnan(o):
            return None
        if isinstance(o, dict):
            return {k: self._clean(v) for k, v in o.items()}
        if isinstance(o, list):
            return [self._clean(v) for v in o]
        return o

class Handler(BaseHTTPRequestHandler):
    def _json(self, data, status=200):
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,PUT,DELETE,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "*")
        self.end_headers()
        try:
            self.wfile.write(json.dumps(data, ensure_ascii=False, indent=2, cls=NanToNullEncoder).encode())
        except:
            self.wfile.write(json.dumps(data, ensure_ascii=False, indent=2).encode())
    
    def _read_body(self):
        length = int(self.headers.get("Content-Length", 0))
        if length == 0: return {}
        body = self.rfile.read(length).decode("utf-8")
        try: return json.loads(body)
        except: return {}
    
    def do_OPTIONS(self):
        self._json({"ok": True})
    
    # ── GET ──────────────────────────────────────
    def do_GET(self):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        
        # /data/today — 获取今日原始数据
        if parsed.path == "/data/today":
            date_str = (qs.get("date") or [date.today().strftime("%Y-%m-%d")])[0]
            code = (qs.get("code") or [None])[0]
            data, err = load_raw(date_str)
            if err:
                return self._json({"error": err}, 404)
            meta = data.pop("_meta", {})
            stocks = data
            total = len(stocks)
            if code:
                stock = stocks.get(code)
                if not stock:
                    return self._json({"error": f"{code} 不在 {date_str} 数据中"}, 404)
                stock["_meta"] = meta
                return self._json(stock)
            # 按持仓过滤
            if "hold" in qs:
                wl = load_watchlist()
                hold_codes = {s["code"] for s in wl["watchlist"]}
                stocks = {k: v for k, v in stocks.items() if k in hold_codes}
            # 分页
            page = int((qs.get("page") or [1])[0])
            size = int((qs.get("size") or [50])[0])
            keys = sorted(stocks.keys())
            start = (page - 1) * size
            end = start + size
            page_keys = keys[start:end]
            return self._json({
                "_meta": {**meta, "total": total, "page": page, "size": size,
                          "returned": len(page_keys)},
                "stocks": {k: stocks[k] for k in page_keys},
                "fields": FIELD_DEFS
            })
        
        # /data/list — 列出可用日期
        if parsed.path == "/data/list":
            files = sorted([f.replace(".json", "") for f in os.listdir(RAW_DIR) if f.endswith(".json")], reverse=True)
            return self._json({"dates": files})
        
        # /data/fields — 字段白名单
        if parsed.path == "/data/fields":
            return self._json({"fields": FIELD_DEFS})
        
        # /data/export — 导出Excel（前端用POST传codes）
        if parsed.path == "/data/export":
            return self._json({"error": "请用 POST 方式导出，传入 codes 参数"}, 405)
        
        # /watchlist — 持仓列表
        if parsed.path == "/watchlist":
            data = load_watchlist()
            return self._json({"codes": [s["code"] for s in data["watchlist"]],
                               "stocks": data["watchlist"]})
        
        # /data/leaders — 行业龙头标签
        if parsed.path == "/data/leaders":
            path = os.path.join(WORKSPACE, "data", "sector_leaders.json")
            if os.path.exists(path):
                with open(path) as f:
                    return self._json(json.load(f))
            return self._json({"leaders": {}, "total_sectors": 0, "total_stocks": 0})
        
        # /score/status — 评分状态
        if parsed.path == "/score/status":
            today = date.today().strftime("%Y-%m-%d")
            score_path = os.path.join(SCORES_DIR, f"{today}.xlsx")
            raw_path = os.path.join(RAW_DIR, f"{today}.json")
            return self._json({
                "raw_exists": os.path.exists(raw_path),
                "raw_size_kb": round(os.path.getsize(raw_path)/1024, 1) if os.path.exists(raw_path) else 0,
                "raw_total": len(json.load(open(raw_path))) - 1 if os.path.exists(raw_path) else 0,
                "score_exists": os.path.exists(score_path),
            })
        
        # / 或 /selestock — 市场全景页面
        if parsed.path == "/" or parsed.path == "/selestock" or parsed.path == "/selestock/":
            return self._serve_file(os.path.join(WORKSPACE, "daily-report-html", "selestock", "index.html"))
        if parsed.path == "/simple.html":
            return self._serve_file(os.path.join(WORKSPACE, "daily-report-html", "selestock", "simple.html"))
        # /debug.html — 诊断工具
        if parsed.path == "/debug.html":
            return self._serve_file(os.path.join(WORKSPACE, "daily-report-html", "selestock", "debug.html"))
        
        self._json({"error": "not found", "paths": ["/", "/selestock", "/data/today","/data/list","/data/fields","/watchlist","/score/status"]}, 404)
    
    def _serve_file(self, path):
        if not os.path.exists(path):
            return self._json({"error": "not found"}, 404)
        ext = os.path.splitext(path)[1]
        mime = {".html": "text/html; charset=utf-8", ".js": "application/javascript",
                ".css": "text/css", ".json": "application/json",
                ".png": "image/png", ".jpg": "image/jpeg", ".svg": "image/svg+xml"}
        self.send_response(200)
        self.send_header("Content-Type", mime.get(ext, "application/octet-stream"))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        with open(path, "rb") as f:
            self.wfile.write(f.read())
    
    # ── POST ─────────────────────────────────────
    def do_POST(self):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        
        # /data/today — 修正某只股票数据
        if parsed.path == "/data/today":
            date_str = (qs.get("date") or [date.today().strftime("%Y-%m-%d")])[0]
            body = self._read_body()
            code = body.get("code") or (qs.get("code") or [""])[0]
            if not code:
                return self._json({"error": "missing code"}, 400)
            data, err = load_raw(date_str)
            if err:
                return self._json({"error": err}, 404)
            if code not in data:
                # 新增股票
                data[code] = {"code": code}
            for k, v in body.items():
                if k != "code":
                    data[code][k] = v
            save_raw(date_str, data)
            return self._json({"status": "ok", "code": code, "message": f"{code} 已更新"})
        
        # /watchlist/add — 加仓
        if parsed.path == "/watchlist/add":
            code = (qs.get("code") or [""])[0]
            name = (qs.get("name") or [""])[0]
            if not code or not name:
                return self._json({"error": "missing code/name"}, 400)
            data = load_watchlist()
            if any(s["code"] == code for s in data["watchlist"]):
                return self._json({"status": "exists", "message": "已在持仓中"})
            sector = qs.get("sector", [get_sector(code)])[0]
            data["watchlist"].append({"code": code, "name": name, "sector": sector})
            save_watchlist(data)
            return self._json({"status": "ok", "message": f"{name} 已加入持仓", "total": len(data["watchlist"])})
        
        # /score/generate — 触发评分计算
        if parsed.path == "/score/generate":
            date_str = (qs.get("date") or [date.today().strftime("%Y-%m-%d")])[0]
            import subprocess
            try:
                r = subprocess.run(["python3", "scripts/calc_score.py", date_str],
                                 capture_output=True, text=True, timeout=120, cwd=WORKSPACE)
                out = r.stdout + r.stderr
                success = "✅" in r.stdout or "[RESULT]" in r.stdout
                return self._json({
                    "status": "ok" if success else "error",
                    "output": out[-1000:],
                    "date": date_str
                })
            except subprocess.TimeoutExpired:
                return self._json({"status": "error", "message": "评分超时(>120s)"}, 504)
            except Exception as e:
                return self._json({"status": "error", "message": str(e)}, 500)
        
        # /data/export — 导出Excel（POST，接收codes数组）
        if parsed.path == "/data/export":
            body = self._read_body()
            codes = body.get("codes", [])
            date_str = (qs.get("date") or [date.today().strftime("%Y-%m-%d")])[0]
            data, err = load_raw(date_str)
            if err:
                return self._json({"error": err}, 404)
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
                from openpyxl.utils import get_column_letter
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "股票数据"
                
                meta = data.pop("_meta", {})
                filed_at = meta.get("fetched_at", "")[:19] if isinstance(meta.get("fetched_at"), str) else "--"
                
                # ── 样式定义 ──
                header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_align = Alignment(horizontal="center", vertical="center")
                cell_font = Font(name="微软雅黑", size=10, color="333333")
                cell_align = Alignment(horizontal="center", vertical="center")
                thin_border = Border(
                    left=Side(style="thin", color="30363d"),
                    right=Side(style="thin", color="30363d"),
                    top=Side(style="thin", color="30363d"),
                    bottom=Side(style="thin", color="30363d")
                )
                up_font = Font(name="微软雅黑", size=10, color="C00000")
                down_font = Font(name="微软雅黑", size=10, color="00B050")
                
                # ── 表头 ──
                headers = ["代码","名称","最新价","涨跌幅%","最高","最低","今开","昨收",
                           "成交额(亿)","流通市值(亿)","市净率","换手率%","市盈率(动)",
                           "量比","市盈率(静)","ROE%","细分行业","主力净流入(万)"]
                ws.append(headers)
                for col_idx, cell in enumerate(ws[1], 1):
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = thin_border
                ws.row_dimensions[1].height = 28
                
                # ── 数据 ──
                # 如果前端传了codes则只导这些，否则导全部
                stock_items = sorted(data.items())
                if codes:
                    code_set = set(codes)
                    stock_items = [(k, v) for k, v in stock_items if k in code_set]
                
                # 隔行背景色
                row_fill_even = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
                
                for row_idx, (code, s) in enumerate(stock_items):
                    row_data = [
                        code,
                        s.get("f14", ""),
                        s.get("f2"), s.get("f3"), s.get("f15"), s.get("f16"),
                        s.get("f17"), s.get("f18"), s.get("f20"), s.get("f21"),
                        s.get("f23"), s.get("f62"), s.get("f115"), s.get("f168"),
                        s.get("f175"), s.get("f184"), s.get("f100"), s.get("f193")
                    ]
                    ws.append(row_data)
                    r = row_idx + 2
                    ws.row_dimensions[r].height = 20
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=r, column=col_idx)
                        cell.border = thin_border
                        cell.alignment = cell_align
                        # 隔行颜色
                        if row_idx % 2 == 0:
                            cell.fill = row_fill_even
                        # 涨跌幅列（第4列）特殊着色
                        if col_idx == 4 and cell.value is not None:
                            try:
                                cell.font = up_font if float(cell.value) > 0 else (down_font if float(cell.value) < 0 else cell_font)
                            except:
                                cell.font = cell_font
                        else:
                            cell.font = cell_font
                
                # ── 列宽 ──
                for col_idx in range(1, len(headers) + 1):
                    max_len = len(headers[col_idx - 1])
                    for row_idx in range(2, len(stock_items) + 2):
                        v = ws.cell(row=row_idx, column=col_idx).value
                        if v is not None:
                            max_len = max(max_len, len(str(v)))
                    ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 22)
                
                # ── 底部信息 ──
                ws.append([])
                info_row = len(stock_items) + 3
                ws.cell(row=info_row, column=1, value=f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(size=9, color="888888")
                ws.cell(row=info_row + 1, column=1, value=f"数据日期: {filed_at}").font = Font(size=9, color="888888")
                ws.cell(row=info_row + 2, column=1, value=f"股票数量: {len(stock_items)}").font = Font(size=9, color="888888")
                
                # ── 冻结首行 ──
                ws.freeze_panes = "A2"
                
                # ── 保存并返回 ──
                out_path = os.path.join(WORKSPACE, "daily-report-html", f"export_{date_str}.xlsx")
                wb.save(out_path)
                
                self.send_response(200)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{os.path.basename(out_path)}"')
                self.send_header("Access-Control-Allow-Origin", "*")
                self.end_headers()
                with open(out_path, "rb") as f:
                    self.wfile.write(f.read())
                os.remove(out_path)
                return
            except Exception as e:
                import traceback
                return self._json({"error": f"导出失败: {e}", "trace": traceback.format_exc()}, 500)
        
        self._json({"error": "not found"}, 404)
    
    # ── DELETE ────────────────────────────────────
    def do_DELETE(self):
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        
        # /data/today — 删除某只股票
        if parsed.path == "/data/today":
            date_str = (qs.get("date") or [date.today().strftime("%Y-%m-%d")])[0]
            code = (qs.get("code") or [""])[0]
            if not code:
                return self._json({"error": "missing code"}, 400)
            data, err = load_raw(date_str)
            if err:
                return self._json({"error": err}, 404)
            if code in data:
                del data[code]
                save_raw(date_str, data)
                return self._json({"status": "ok", "message": f"{code} 已删除"})
            return self._json({"error": f"{code} 不存在"}, 404)
        
        # /watchlist/remove — 减仓
        if parsed.path == "/watchlist/remove":
            code = (qs.get("code") or [""])[0]
            if not code:
                return self._json({"error": "missing code"}, 400)
            data = load_watchlist()
            before = len(data["watchlist"])
            data["watchlist"] = [s for s in data["watchlist"] if s["code"] != code]
            save_watchlist(data)
            return self._json({"status": "ok", "message": "已移除", "total": len(data["watchlist"])})
        
        self._json({"error": "not found"}, 404)
    
    def log_message(self, format, *args):
        sys.stderr.write(f"[DataService] {format % args}\n")

# ─── 启动 ────────────────────────────────────────
def stop_existing():
    """杀掉同端口旧进程"""
    import subprocess
    try:
        r = subprocess.run(["lsof", "-ti", ":18790"], capture_output=True, text=True, timeout=5)
        for pid in r.stdout.strip().split("\n"):
            if pid:
                os.system(f"kill -9 {pid} 2>/dev/null")
                print(f"  🧹 终止旧进程: {pid}")
    except:
        pass

if __name__ == "__main__":
    port = 18790
    print(f"🚀 DataService starting on http://127.0.0.1:{port}")
    print(f"   📊 GET  /data/today        今日数据（支持 ?code=&page=&hold=1）")
    print(f"   📜 GET  /data/list          可用日期列表")
    print(f"   📋 GET  /data/fields        字段白名单")
    print(f"   📝 POST /data/today         修正/新增股票数据")
    print(f"   🗑️  DELETE /data/today       删除股票数据")
    print(f"   📌 GET  /watchlist          持仓列表")
    print(f"   ➕ POST /watchlist/add      加仓")
    print(f"   ➖ DELETE /watchlist/remove 减仓")
    print(f"   🔄 POST /score/generate     触发评分")
    print(f"   📊 GET  /score/status       评分状态")
    print(f"   RAW: {RAW_DIR}")
    print(f"   WATCH: {WATCH_CONFIG}")
    
    server = HTTPServer(("127.0.0.1", port), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n🛑 服务停止")
        server.server_close()
