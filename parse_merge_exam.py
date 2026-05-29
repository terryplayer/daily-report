#!/usr/bin/env python3
"""Parse multiple AI exam HTML files, merge, deduplicate, and update the existing Excel."""

import re
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# File paths
HTML_FILES = {
    "唐友权": "/Users/shisan/.openclaw/media/inbound/AI编程能力认证考试_1---64564459-ed60-475e-852e-ad5d5deb8b46.html",
    "李道泽": "/Users/shisan/.openclaw/media/inbound/AI编程能力认证考试1---b3b627b7-db0a-4fdf-a9c4-fba01dfb3322.html",
    "田科": "/Users/shisan/.openclaw/media/inbound/AI编程能力认证考试---4dd372d6-9d79-45bc-a972-0a4991ab401f.html",
}
OLD_XLSX = "/Users/shisan/.openclaw/workspace/AI编程能力认证考试_题目解析.xlsx"
NEW_XLSX = "/Users/shisan/.openclaw/workspace/AI编程能力认证考试_题目解析_合并.xlsx"

def parse_exam_html(filepath, student_name):
    """Parse exam HTML and return list of question dicts."""
    with open(filepath, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f.read(), 'html.parser')

    rv_area = soup.find('div', id='rvArea')
    if not rv_area:
        print(f"  ⚠ No rvArea found for {student_name}, trying qArea...")
        q_area = soup.find('div', id='qArea')
        if q_area:
            return parse_qarea(q_area, student_name, soup)
        return []
    
    print(f"  📝 Using rvArea (review page)")
    return parse_rvarea(rv_area, student_name)

def parse_qarea(q_area, student_name, soup):
    """Parse exam-in-progress page (qArea format with watermark answers)."""
    # Extract watermark answers
    wm_layer = soup.find('div', id='wmLayer')
    wm_answers = {}
    if wm_layer:
        for span in wm_layer.find_all('span'):
            m = re.match(r'Q(\d+):([A-Z]+)', span.get_text(strip=True))
            if m:
                qnum = int(m.group(1))
                answer = m.group(2)
                wm_answers[qnum] = answer
    
    # Get user name
    user_elem = soup.find('b', id='examUser')
    user_name = user_elem.get_text(strip=True) if user_elem else student_name
    
    questions = []
    qcards = q_area.find_all('div', class_='qcard', recursive=False)
    
    for qcard in qcards:
        try:
            head = qcard.find('div', class_='qcard-head')
            if not head:
                continue
            spans = head.find_all('span')
            if len(spans) < 4:
                continue
            
            qid_tag = spans[0].get_text(strip=True)
            level = spans[1].get_text(strip=True)
            qtype_raw = spans[2].get_text(strip=True)
            score = spans[3].get_text(strip=True)
            
            # Clean type (remove bold tags for 多选)
            qtype = qtype_raw.replace('多选', '多选').replace('单选', '单选').replace('判断', '判断')
            
            text_div = qcard.find('div', class_='qcard-text')
            if not text_div:
                continue
            full_text = text_div.get_text(strip=True)
            m = re.match(r'(\d+)\.\s*(.*)', full_text)
            qnum = m.group(1) if m else str(len(questions) + 1)
            qtext = m.group(2) if m else full_text
            
            # Extract options and user's selection
            opts_div = qcard.find('div', class_='opts')
            opt_items = opts_div.find_all('div', class_=re.compile(r'opt')) if opts_div else []
            
            labels = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            options_data = []
            user_selections = []
            
            for idx, opt in enumerate(opt_items):
                classes = opt.get('class', [])
                is_selected = 'sel' in classes
                
                for dcy in opt.find_all('span', class_='dcy'):
                    dcy.decompose()
                opt_text = opt.get_text(strip=True)
                
                label = labels[idx] if idx < len(labels) else chr(65 + idx)
                options_data.append({
                    'label': label,
                    'text': opt_text,
                    'is_selected': is_selected
                })
                
                if is_selected:
                    user_selections.append(label)
            
            # Determine correct answer from watermark
            qnum_int = int(qnum)
            if qnum_int in wm_answers:
                correct_str = wm_answers[qnum_int]
            else:
                correct_str = ''
            
            options_text = '\n'.join([f"{o['label']}. {o['text']}" for o in options_data])
            
            # Compare user selection with correct answer
            user_sel_str = ', '.join(user_selections) if user_selections else ''
            
            if qtype == '多选':
                user_sel_set = set(user_selections)
                correct_set = set(correct_str.replace(', ', ',').split(',')) if correct_str else set()
                if user_sel_set == correct_set:
                    result_status = '正确'
                elif user_sel_set:
                    result_status = '部分正确' if user_sel_set & correct_set else '错误'
                else:
                    result_status = '未作答'
            else:
                # For 单选/判断, compare first char
                if user_sel_str and correct_str:
                    # Normalize: for 判断, "对" vs "错"
                    if qtype == '判断':
                        # In watermark, might be A/B mapped to 对/错
                        if user_sel_str == correct_str:
                            result_status = '正确'
                        else:
                            result_status = '错误'
                    else:
                        # 单选: watermark answer maps to option letter
                        if user_sel_str == correct_str:
                            result_status = '正确'
                        else:
                            result_status = '错误'
                elif user_sel_str:
                    result_status = '已作答'
                else:
                    result_status = '未作答'
            
            # For 判断, we need to check the actual text
            if qtype == '判断' and len(options_data) >= 2:
                # Options are 对/错
                correct_letter = correct_str
                if correct_letter == 'A' or correct_letter == '对':
                    correct_letter_display = '对（正确）'
                    actual_correct = 'A'
                elif correct_letter == 'B' or correct_letter == '错':
                    correct_letter_display = '错（错误）'
                    actual_correct = 'B'
                else:
                    actual_correct = correct_letter
            
            questions.append({
                'student': user_name,
                'qid_tag': qid_tag,
                'level': level,
                'type': qtype,
                'score': score,
                'qnum': qnum,
                'question_text': qtext,
                'options': options_text,
                'correct_answer': correct_str,
                'user_answer': user_sel_str,
                'result': result_status,
                'explanation': '',
                'reference': '',
                'source_file': os.path.basename(filepath),
                'full_question': full_text
            })
            
        except Exception as e:
            print(f"  Error parsing question: {e}")
            continue
    
    return questions

def parse_rvarea(rv_area, student_name):
    """Parse review page (rvArea format with marked answers)."""
    questions = []
    qcards = rv_area.find_all('div', class_=re.compile(r'qcard'))
    
    for qcard in qcards:
        try:
            head = qcard.find('div', class_='qcard-head')
            if not head:
                continue
            spans = head.find_all('span')
            if len(spans) < 5:
                continue
            
            qid_tag = spans[0].get_text(strip=True)
            level = spans[1].get_text(strip=True)
            qtype = spans[2].get_text(strip=True)
            score = spans[3].get_text(strip=True)
            result_tag = spans[4].get_text(strip=True)
            
            text_div = qcard.find('div', class_='qcard-text')
            if not text_div:
                continue
            full_text = text_div.get_text(strip=True)
            m = re.match(r'(\d+)\.\s*(.*)', full_text)
            qnum = m.group(1) if m else str(len(questions) + 1)
            qtext = m.group(2) if m else full_text
            
            opts_div = qcard.find('div', class_='opts')
            opt_items = opts_div.find_all('div', class_=re.compile(r'opt')) if opts_div else []
            
            labels = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
            options_data = []
            user_selections = []
            
            for idx, opt in enumerate(opt_items):
                classes = opt.get('class', [])
                is_correct = 'rv-right' in classes
                is_wrong_sel = 'rv-wrong-sel' in classes
                is_missed = 'rv-missed' in classes
                
                for dcy in opt.find_all('span', class_='dcy'):
                    dcy.decompose()
                opt_text = opt.get_text(strip=True)
                
                label = labels[idx] if idx < len(labels) else chr(65 + idx)
                options_data.append({
                    'label': label,
                    'text': opt_text,
                    'is_correct': is_correct,
                    'is_selected': is_wrong_sel or (is_correct and 'rv-missed' not in classes),
                })
                
                if is_wrong_sel:
                    user_selections.append(label)
                elif is_correct and not is_missed and 'rv-right' in classes:
                    pass  # correct answer, not necessarily selected
            
            correct_answers = [o['label'] for o in options_data if o['is_correct']]
            correct_str = ', '.join(correct_answers) if correct_answers else ''
            
            options_text = '\n'.join([f"{o['label']}. {o['text']}" for o in options_data])
            
            explain_div = qcard.find('div', class_='qcard-explain')
            explanation = ''
            reference = ''
            if explain_div:
                et = explain_div.find('div', class_='qcard-explain-text')
                if et:
                    explanation = et.get_text(strip=True)
                src = explain_div.find('div', class_='qcard-source')
                if src:
                    reference = src.get_text(strip=True)
            
            if 'correct' in result_tag:
                result_status = '正确'
            elif 'partial' in result_tag:
                result_status = '部分正确'
            elif 'wrong' in result_tag:
                result_status = '错误'
            else:
                result_status = result_tag.strip()
            
            questions.append({
                'student': student_name,
                'qid_tag': qid_tag,
                'level': level,
                'type': qtype,
                'score': score,
                'qnum': qnum,
                'question_text': qtext,
                'options': options_text,
                'correct_answer': correct_str,
                'user_answer': ', '.join(user_selections) if user_selections else '',
                'result': result_status,
                'explanation': explanation,
                'reference': reference,
                'source_file': os.path.basename(HTML_FILES.get(student_name, '')),
                'full_question': full_text
            })
            
        except Exception as e:
            print(f"  Error parsing question: {e}")
            continue
    
    return questions

# ────── Main ──────

# Parse all HTML files
all_questions = []
for student, path in HTML_FILES.items():
    print(f"\n📄 Parsing {student}'s exam...")
    qs = parse_exam_html(path, student)
    print(f"   → {len(qs)} questions extracted")
    all_questions.extend(qs)

print(f"\n📊 Total: {len(all_questions)} questions from all files")

# Deduplicate by question text (exact match)
seen_texts = {}
deduped = []
dups_removed = 0

for q in all_questions:
    key = q['question_text'].strip()
    if key in seen_texts:
        # Keep the one with more info (has explanation)
        existing = seen_texts[key]
        if q['explanation'] and not existing['explanation']:
            # Replace with richer version
            deduped.remove(existing)
            deduped.append(q)
            seen_texts[key] = q
        dups_removed += 1
    else:
        seen_texts[key] = q
        deduped.append(q)

print(f"   → {dups_removed} duplicates removed")
print(f"   → {len(deduped)} unique questions remaining")

# ===== CREATE MERGED EXCEL =====
wb = Workbook()

header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='00B4D8', end_color='00B4D8', fill_type='solid')
header_fill2 = PatternFill(start_color='6366F1', end_color='6366F1', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
pass_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
fail_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

def style_header(ws, headers, fill=None):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill or header_fill
        cell.alignment = header_align
        cell.border = thin_border

def style_cell(ws, row, col, value, font=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or Font(name='微软雅黑', size=10)
    cell.alignment = align or Alignment(vertical='top', wrap_text=True)
    cell.border = thin_border
    return cell

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w

# ── Sheet 1: 合并题目列表 ──
ws1 = wb.active
ws1.title = "合并题目列表"

headers1 = ['序号', '题目编号', '等级', '题型', '分值', '题目内容', '选项', '正确答案', '答题结果', '解析', '参考来源', '考生']
style_header(ws1, headers1)

for row_idx, q in enumerate(deduped, 2):
    data = [
        q['qnum'], q['qid_tag'], q['level'], q['type'], q['score'],
        q['question_text'], q['options'], q['correct_answer'], q['result'],
        q['explanation'], q['reference'], q['student']
    ]
    for col_idx, value in enumerate(data, 1):
        cell = style_cell(ws1, row_idx, col_idx, value)
    
    # Color result
    result_cell = ws1.cell(row=row_idx, column=9)
    if '正确' in q['result']:
        result_cell.font = Font(name='微软雅黑', size=10, color='00B050', bold=True)
    elif '部分' in q['result']:
        result_cell.font = Font(name='微软雅黑', size=10, color='FFC000', bold=True)
    elif '错误' in q['result']:
        result_cell.font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)

set_col_widths(ws1, [6, 10, 7, 8, 6, 40, 55, 12, 10, 55, 30, 10])
ws1.freeze_panes = 'A2'

# ── Sheet 2: 原始数据（不分卷） ──
ws2 = wb.create_sheet("所有题目（含重复）")

headers2 = ['序号', '题目编号', '等级', '题型', '分值', '题目内容', '选项', '正确答案', '考生', '答题结果', '解析', '参考来源']
style_header(ws2, headers2, header_fill2)

for row_idx, q in enumerate(all_questions, 2):
    data = [
        q['qnum'], q['qid_tag'], q['level'], q['type'], q['score'],
        q['question_text'], q['options'], q['correct_answer'], q['student'], q['result'],
        q['explanation'], q['reference']
    ]
    for col_idx, value in enumerate(data, 1):
        cell = style_cell(ws2, row_idx, col_idx, value)

set_col_widths(ws2, [6, 10, 7, 8, 6, 40, 55, 12, 10, 10, 55, 30])
ws2.freeze_panes = 'A2'

# ── Sheet 3: 标准答案汇总 ──
ws3 = wb.create_sheet("标准答案")

headers3 = ['题目编号', '等级', '题型', '正确答案', '题目摘要']
style_header(ws3, headers3)

for row_idx, q in enumerate(deduped, 2):
    summary = q['question_text'][:60] + ('...' if len(q['question_text']) > 60 else '')
    data = [q['qid_tag'], q['level'], q['type'], q['correct_answer'], summary]
    for col_idx, value in enumerate(data, 1):
        cell = style_cell(ws3, row_idx, col_idx, value)
    ans_cell = ws3.cell(row=row_idx, column=4)
    ans_cell.font = Font(name='微软雅黑', size=10, color='00B050', bold=True)

set_col_widths(ws3, [12, 8, 10, 15, 55])
ws3.freeze_panes = 'A2'

# ── Sheet 4: 考试信息 ──
ws4 = wb.create_sheet("考试信息")

ws4.cell(row=1, column=1, value='考试信息汇总').font = Font(name='微软雅黑', bold=True, size=14)

info_data = [
    ('', ''),
    ('文件来源', ''),
]
for student, path in HTML_FILES.items():
    info_data.append(('', ''))
    # Extract student info from file
    with open(path, 'r', encoding='utf-8') as f:
        html_content = f.read()
    
    # Try to get user name
    nm = re.search(r'<b id="(?:homeUser|examUser)">([^<]+)</b>', html_content)
    user_name = nm.group(1) if nm else student
    
    # Try to get email
    em = re.search(r'<div class="value" id="hiEmail">([^<]+)</div>', html_content)
    email = em.group(1) if em else '-'
    
    # Count questions
    rv_soup = BeautifulSoup(html_content, 'html.parser')
    rv = rv_soup.find('div', id='rvArea')
    q_count = len(rv.find_all('div', class_=re.compile(r'qcard'))) if rv else 0
    
    info_data.append((f'📋 {user_name}', f'共 {q_count} 题'))
    info_data.append(('  邮箱', email))
    info_data.append(('  文件名', os.path.basename(path)))

row = 3
for k, v in info_data:
    if k:
        style_cell(ws4, row, 1, k, Font(name='微软雅黑', bold=bool(k.startswith('📋')), size=10 if not k.startswith('📋') else 11))
        style_cell(ws4, row, 2, v)
    row += 1

ws4.column_dimensions['A'].width = 20
ws4.column_dimensions['B'].width = 50

# ── Sheet 5: 错题集 ──
ws5 = wb.create_sheet("错题集")

wrong_qs = [q for q in deduped if '错误' in q['result'] or '部分' in q['result']]
headers5 = ['序号', '题目编号', '等级', '题型', '题目内容', '选项', '正确答案', '你的答案', '解析', '参考来源', '考生']
style_header(ws5, headers5)

for row_idx, q in enumerate(wrong_qs, 2):
    data = [
        q['qnum'], q['qid_tag'], q['level'], q['type'],
        q['question_text'], q['options'], q['correct_answer'], q['user_answer'],
        q['explanation'], q['reference'], q['student']
    ]
    for col_idx, value in enumerate(data, 1):
        cell = style_cell(ws5, row_idx, col_idx, value)

if wrong_qs:
    set_col_widths(ws5, [6, 10, 7, 8, 40, 55, 12, 12, 55, 30, 10])
else:
    style_cell(ws5, 2, 1, '🎉 全部正确，没有错题！', Font(name='微软雅黑', size=12, color='00B050'))
    ws5.merge_cells('A2:K2')

ws5.freeze_panes = 'A2'

# Save
wb.save(NEW_XLSX)
print(f"\n✅ Merged Excel saved to: {NEW_XLSX}")
print(f"   Sheet 1: 合并题目列表 ({len(deduped)} 题去重后)")
print(f"   Sheet 2: 所有题目（含重复） ({len(all_questions)} 题)")
print(f"   Sheet 3: 标准答案")
print(f"   Sheet 4: 考试信息")
print(f"   Sheet 5: 错题集 ({len(wrong_qs)} 题)")
